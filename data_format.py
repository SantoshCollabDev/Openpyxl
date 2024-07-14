import openpyxl
from openpyxl.styles  import numbers

wb = openpyxl.load_workbook("example.xlsx")

ws = wb['score']

ws['C4'] = '11/11/20'
ws['C4'].number_format = numbers.FORMAT_DATE_DDMMYY

ws['D4'] = 20
ws['D4'].number_format = numbers.FORMAT_NUMBER

ws['E4'] = 'Beginner'
ws['E4'].number_format = numbers.FORMAT_TEXT


wb.save("example.xlsx")

