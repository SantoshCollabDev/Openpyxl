import openpyxl
from openpyxl.styles import Font
import openpyxl.workbook

master_data = openpyxl.load_workbook("Master_Sheet.xlsx")
daily_data = openpyxl.load_workbook("Daily_Sheet.xlsx")

master_sheet = master_data['mdata']
daily_sheet = daily_data['ddata']

# Get Row Count of Daily Sheet

is_data = True
daily_row_count = 1

while is_data:
  daily_row_count += 1
  data = daily_sheet.cell(row=daily_row_count,column=1).value
  if data == None:
    is_data = False

# print(daily_row_count)


# Get Row Count of Master Sheet

is_data = True
master_row_count = 1

while is_data:
  master_row_count += 1
  data = master_sheet.cell(row=master_row_count,column=1).value
  if data == None:
    is_data = False

# print(master_row_count)


# Get data from daily sheet
# extract the data ---> store into list of dictionaries
todays_data = []

for i in range(1,daily_row_count):
  row_data = {}
  row_data['id'] = daily_sheet.cell(row=i, column=1).value
  row_data['todays_purchase'] = daily_sheet.cell(row=i, column=2).value
  row_data['todays_rewards'] = daily_sheet.cell(row=i, column=3).value
  todays_data.append(row_data)

# print(todays_data)

# WRITE daily sheet data into MASTER SHEET
# 1. find a row using id i.e. key column
# 2. Add todays purchse to Total_Purchase
# 3. Add todays reward to Total_Rewards

# 1. Iterate over master sheet and find rows that need update

for i in range(2,master_row_count):
  id = master_sheet.cell(row=i, column=1).value
  for row in todays_data:
    if row['id'] == id:
      todays_purchase = int(row['todays_purchase'])
      todays_rewards = int(row['todays_rewards'])

      # Get data from master sheet
      total_purchase = master_sheet.cell(row=i,column=6).value
      total_reward = master_sheet.cell(row=i,column=7).value

      # Add values of todays data into total data
      new_total_purchase = total_purchase + todays_purchase
      new_total_reward = total_reward + todays_rewards


      master_sheet.cell(row=i,column=6).value = new_total_purchase
      master_sheet.cell(row=i,column=7).value = new_total_reward

# create a blank object -- to write data abd save as xl
daily_report = openpyxl.Workbook()
ws = daily_report.active

# get headers
is_data = True
column_count = 1
header_values = []

while is_data:
  column_count += 1
  data = master_sheet.cell(row=1,column=column_count).value
  if data != None:
    header_values.append(data)
  else:
    is_data = False

# print(header_values)

header_style = Font(name="Times New Roman", 
                    size=12,
                    bold = True)

for i, col_name in enumerate(header_values):
    # print(i,col_name)
    col_index = i + 1
    ws.cell(row=1,column=col_index).value = col_name
    ws.cell(row=1,column=col_index).font = header_style

# grab the ids of the daily data
ids = []
for data in todays_data:
  ids.append(data['id'])

ids.pop(0)
print(ids)

# get the data from master to write to report

final_data = []
for i in range(2,master_row_count):
  id = master_sheet.cell(row=i, column=1).value
  if id in ids:
    lst = []
    for  j in  range(2,8): # 7 columns
      lst.append(master_sheet.cell(row=i, column=j).value)
    final_data.append(lst)    
# print(final_data)

for data in final_data:
  ws.append(data)

daily_report.save("daily_report_send.xlsx")

master_data.save("Master_Sheet.xlsx")
