import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

ws = wb['balance']

# iterate through rows
rows = ws.iter_rows(min_row=1, max_row=7, 
                    min_col=1, max_col=2)
# print(rows)

names = []
balance = []

# for row in rows:
for a,b in rows:
  # print(a.value, b.value)
  names.append(a.value)
  balance.append(b.value)

# print(names)
# print(balance)

# ITERATE THROUGH COLUMNS
# min_row = 1 & min_col = 1 optional if we are always starting from 1
# columns = ws.iter_cols(min_row=1, max_row=7, 
#                        min_col=1, max_col=2)
# to read all cols and rows simply use -- ws.iter_cols() w/o any paratmers

columns = ws.iter_cols()
# print(columns)

# for col in columns:
#   print(col)

# get access all rows
rows = list(ws.rows)
print(rows)
# get access all columns
columns = list(ws.columns)
print(columns)


wb.save("example.xlsx")
