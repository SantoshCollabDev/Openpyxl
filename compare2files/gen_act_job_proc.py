import openpyxl

def load_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return sheet

def get_column_data(sheet, column_index):
    column_data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        column_data[row[column_index]] = row

    print(column_data)    
    return column_data

def write_matching_records(sheet1_data, sheet2_data, output_file):
    workbook = openpyxl.Workbook()
    output_sheet = workbook.active

    # Write headers (assuming both sheets have the same headers)
    # headers = ["Job Name", "Proc"]  # Add your actual headers here
    # output_sheet.append(headers)

    # Write matching records
    # print(sheet1_data.keys())
    # print(sheet2_data.keys())
    for key in sheet1_data.keys() & sheet2_data.keys():
        output_sheet.append(sheet1_data[key])

    workbook.save(output_file)

# Load the two Excel files
sheet1 = load_excel('Job_Proc.xlsx')
sheet2 = load_excel('Active_Job.xlsx')

# Get data from the first column (index 0)
sheet1_data = get_column_data(sheet1, 0)
sheet2_data = get_column_data(sheet2, 0)

# Write matching records to a new Excel file
write_matching_records(sheet1_data, sheet2_data, 'matching_records.xlsx')



# import openpyxl

# wb1 = openpyxl.load_workbook("Job_Proc.xlsx")
# wb2 = openpyxl.load_workbook("Active_Job.xlsx")
# wb3 = openpyxl.load_workbook("Active_Job_mapping.xlsx")

# jp = wb1['Sheet1']
# aj = wb2['Sheet1']
# ml = wb3.active

# match_jobs = []

# for rwaj in aj.iter_rows():
#     # print("outer loop")
#     for rwjp in jp.iter_rows():
#         # print("inner loop")
#         # print(rwjp[0].value)
#         if rwjp[0].value == rwaj[0].value:
#             # print("match_found")
#             # print(rwaj[0].value)
#             # match_jobs.append(rwjp)
#             ml.append(rwjp)
#     # print(match_jobs)
#     # print("--------------------------")
# # print(match_jobs)
# # for row in match_jobs:
# #   ml.append(row)

# wb3.save("Active_Job_mapping.xlsx")