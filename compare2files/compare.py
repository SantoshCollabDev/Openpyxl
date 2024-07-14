import openpyxl
from openpyxl.styles import PatternFill

data_file1 = openpyxl.load_workbook("Master_Copy.xlsx")
data_file2 = openpyxl.load_workbook("Master_Sheet.xlsx")


fill_style = PatternFill(start_color="FDD835",
                         end_color="FDD835",
                         fill_type="solid")

data_sheet1 = data_file1['cdata']
data_sheet2 = data_file2['mdata']

for row in data_sheet1.iter_rows():
  for cell in row:
    current_cell_value = cell.value
    cell_location = cell.coordinate

    if current_cell_value != data_sheet2[cell_location].value:
      cell.fill = fill_style

data_file1.save("compared_file.xlsx")
