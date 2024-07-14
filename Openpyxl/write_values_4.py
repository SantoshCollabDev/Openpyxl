import openpyxl

wb = openpyxl.load_workbook("Example.xlsx")

ws = wb['Balance']

# Create a COLUMN C whose values would be double the values of balance i.e. COLUMN B

#  ws.max_row  <----- give you the number of rows in a worksheet.

ws.cell(1,3).value = "Double Balance"

for r in range(2, ws.max_row + 1):  # for each row in excel starting from 2 by leaving header i.e. row 1
  b_col = ws.cell(row=r, column=2).value  # col=2 is fixed since that contains balance to be doubled
  c_value = b_col * 2
  ws.cell(row=r, column=3).value = c_value
  

wb.save("Example.xlsx")


