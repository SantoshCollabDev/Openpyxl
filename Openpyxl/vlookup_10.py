import openpyxl

mb = openpyxl.load_workbook("master_data.xlsx")
db = openpyxl.load_workbook("daily_data.xlsx")

master_sheet = mb['master_data']
daily_sheet  = db['daily_data']

# iter_rows() -->  TO ITERATE THROUGH THE ROWS OF AN EXCEL
# .row  --> TO GET THE ROW NUMBER OF THE CURRENT ROW IN PROCESSING


for i in daily_sheet.iter_rows():
  for j in master_sheet.iter_rows():
    if j[0].value == i[0].value:
      # Append the columns B, C and D to the daily sheet
      # get the row number at which value should be appended
      row_number = i[0].row
      daily_sheet.cell(row=row_number, column=4).value = j[1].value
      daily_sheet.cell(row=row_number, column=5).value = j[2].value
      daily_sheet.cell(row=row_number, column=6).value = j[3].value


db.save("daily_data.xlsx")