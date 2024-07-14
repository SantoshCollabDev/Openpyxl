import openpyxl

# to work with excel formatting 
# OPENPYXL provides 2 styling modules ----> Font, Color
from openpyxl.styles import Font, Color
 
wb = openpyxl.load_workbook("Example.xlsx")

ws = wb["Balance"]


# define the font styling using the ---> Font class of openpyxl

# ****************************************************************
# REQUIREMENT 1:  change the font style of A4 
# https://www.rapidtables.com/web/color/RGB_Color.html
# ****************************************************************

font_style = Font(name="Arial Black", size=14, color="1807FF")  # define & save the font style

ws['A4'].font = font_style   # APPLY font style on the cell

# ****************************************************************
# REQUIREMENT 2:  change the font style of entire column 3 i.e. couble bonus
# https://www.rapidtables.com/web/color/RGB_Color.html
# ****************************************************************

column_style = Font(name="Arial Black", size=12, color="FF07E6", bold=True, italic=True, underline="none")

for r  in range(2, ws.max_row + 1):
  ws.cell(row=r, column=3).font = column_style


wb.save("Example.xlsx")

