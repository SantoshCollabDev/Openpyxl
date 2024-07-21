import openpyxl
from openpyxl.styles import PatternFill

file1 = openpyxl.load_workbook("Data_file1.xlsx")
file2 = openpyxl.load_workbook("Data_file2.xlsx")

file1_data = file1["master_data"]
file2_data = file2["master_data"]

fill_style = PatternFill(start_color="E60D0D", end_color="E60D0D", fill_type="solid")

# cell.coordinate --- GET THE LOCATION OF CURRENT CELL
for row in file1_data.iter_rows():
  for cell in row:
    current_cell_value = cell.value
    cell_location = cell.coordinate

# cell.fill --- TO SET FILL STYLE
    if current_cell_value != file2_data[cell_location].value:
      cell.fill = fill_style

file1.save("compared_file.xlsx")

