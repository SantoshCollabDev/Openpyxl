import openpyxl

wb = openpyxl.load_workbook("Example.xlsx")

ws = wb['Balance']


# *****************************************
# Accessing values using iter_rows()
# *****************************************
# read first 5 rows of Balance sheet
rows = ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2)

print(rows) # gives a generator objct as shown below
# <generator object Worksheet._cells_by_row at 0x0000028663B7D000>

# to access values
for r,c in rows:
  print(r.value, c.value)


# *****************************************
# Accessing values using iter_cols()
# *****************************************
# NO NEED TO SPECIFY MIN OR MAX PARAMETERS WHEN HAVE TO 
# READ ALL THE ROWS AND COLUMNS
cols = ws.iter_cols()  
print(cols)

for tup in cols:
  for val in tup:
    print(val.value)


# *********************************************************************
# NOTE: ALWAYS USE MIN/MAX paratmeters with iter_rows() or iter_cols();
# We have other mechanisms to access rows and cols without
# specifying min/max ranges
# ********

# WS.ROWS   --- to access rows
all_rows = ws.rows

for row,col in all_rows:
  print(row.value, col.value)


# WS.COLS   --- to access COLS
all_cols = ws.columns

for tup in all_cols:
  for val in tup:
    print(val.value)