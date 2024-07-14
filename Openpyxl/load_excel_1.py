import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

# print all sheet names in a excel
print(wb.sheetnames)

#select a specific worksheet from workbook
ws1 = wb['Score']
print(ws1)

ws2 = wb['Balance']
print(ws2)

# Add a new sheet at specificed index
wb.create_sheet("NewSheet",index=0)  # add as 1st sheet
wb.save("example.xlsx")  #always save
