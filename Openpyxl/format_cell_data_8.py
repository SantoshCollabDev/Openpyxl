import openpyxl
from openpyxl.styles import numbers

wb = openpyxl.load_workbook("Example.xlsx")
ws = wb['Score']

ws['C4'] = '2024-07-15'
ws['C4'].number_format = numbers.FORMAT_DATE_DATETIME

ws['D4'] = 20
ws['D4'].number_format = numbers.FORMAT_NUMBER

ws['E4'] = 'Beginner'
ws['E4'].number_format = numbers.FORMAT_TEXT

wb.save("Example.xlsx")