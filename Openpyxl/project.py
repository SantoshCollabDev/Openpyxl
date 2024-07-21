'''
Requirement:

1. Add daily transaction data in master data file
  a. Add Today's purchase in Total purchases
  b. Add Today's reward in Lifetime Reward Balance

2. Create a new daily report
  a. Create a new report that contain FirstName, LastName, Email, Gender, Total Purchases and Lifetime Reward Balance Columns
  b. Headers of this should be 12pt, Times New Roman, Bold

'''

import openpyxl
import openpyxl.workbook
from openpyxl.styles import Font

master_data = openpyxl.load_workbook("Master_Sheet.xlsx")
daily_data  = openpyxl.load_workbook("Daily_Sheet.xlsx")

master_sheet = master_data['mdata']
daily_sheet  = daily_data['ddata']


#==========================================
# Requirement: 1
#==========================================

# get ROW COUNT for daily sheet
is_data = True
daily_row_count = 1  # in openpyxl indexing starts from 1

while is_data:
  data = daily_sheet.cell(row=daily_row_count, column=1).value
  # print(data)
  if data == None:
    is_data = False
  else:
    daily_row_count += 1
# print("daily_row_count :",daily_row_count-1)


# get ROW COUNT for master sheet
is_data = True
master_row_count = 1  # in openpyxl indexing starts from 1

while is_data:
  data = master_sheet.cell(row=master_row_count, column=1).value
  # print(data)
  if data == None:
    is_data = False
  else:
    master_row_count += 1
# print("master_row_count :",master_row_count-1)


# get data from daily sheet
# Extract data ---> store it into list of dictionaries
todays_data = []

for i in range(1, daily_row_count):
  row_data = {}
  row_data['id']               = daily_sheet.cell(row=i,column=1).value
  row_data['todays_purchases'] = daily_sheet.cell(row=i,column=2).value
  row_data['todays_reward']    = daily_sheet.cell(row=i,column=3).value
  todays_data.append(row_data)

# print(todays_data)
# {'id': 'ID', 'Todays Pruchases': 'Todays Pruchases', 'Todays Reward': 'Todays Reward'}


# write daily sheet data into master excel sheet
# Find row using the ID
# Go to total purchase cell + todays purchase
# Go to total reward balance + todays reward

# Iterate over master sheet to find rows need to be updated

for i in range(2, master_row_count):
  # grab id for each row
  id = master_sheet.cell(row=i, column=1).value
  # check if this id exists in daily data
  for row in todays_data:
    if row['id'] == id:
      # grab todays purchse and todays reward
      todays_purchase = row['todays_purchases']
      todays_reward = row['todays_reward']

      # get data from master sheet
      total_purchase = master_sheet.cell(row=i, column=6).value
      total_reward = master_sheet.cell(row=i, column=7).value

      # Add values of todays purchase into total purchases
      new_total_purchase = total_purchase + todays_purchase
      new_total_reward =  total_reward + todays_reward

      master_sheet.cell(row=i, column=6).value = new_total_purchase
      master_sheet.cell(row=i, column=7).value = new_total_reward

master_data.save('Master_Sheet.xlsx')

#==========================================
# Requirement: 2
'''2. Create a new daily report
  a. Create a new report that contain FirstName, LastName, Email, Gender, Total Purchases and Lifetime Reward Balance Columns
  b. Headers of this should be 12pt, Times New Roman, Bold'''
#==========================================

# create a blank work book object

daily_report = openpyxl.Workbook()
ws = daily_report.active

# Get headers 
is_data = True
column_count = 2
header_values = []

while is_data:
  # column_count += 1  # ignore id column and start from column 2 since we dont need id in final report
  data = master_sheet.cell(row=1, column=column_count).value
  # print(data)
  if data != None:
    header_values.append(data)
  else:
    is_data = False
  column_count += 1

# print(header_values)

# define font styles for header as per requirement

header_style = Font(name="Times New Roman",
                    size=12,
                    bold=True)

# iterate over the header values and add them in new excel sheet

for i, col_name in enumerate(header_values):
  print(i, col_name)
  col_index = i + 1
  ws.cell(row=1, column=col_index).value = col_name
  ws.cell(row=1, column=col_index).font = header_style

# grab the data from current master data sheet and append to new report
# first get id's of todays data from daily sheet
# daily data is avaialble in ---> todays_data[]

IDs = []
for data in todays_data:
  IDs.append(data['id'])
IDs.pop(0)  # pass the index of the value to pop
# print(IDs)

# now get todays data from master sheet and append to daily report

final_data = []
for i in range(2,master_row_count):
  id = master_sheet.cell(row=i,column=1).value
  if id in IDs:
    lst = []
    for j in range(2, 8):  # we dont need ID column so start with 2
      lst.append(master_sheet.cell(row=i, column=j).value)
    final_data.append(lst)

print(final_data)

# write data into excel report
for data in final_data:
  ws.append(data)

daily_report.save("daily_report_send.xlsx")

