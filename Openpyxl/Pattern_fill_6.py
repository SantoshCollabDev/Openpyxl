import openpyxl

from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("Example.xlsx")

ws = wb['Balance']

# Define the pattern/color fill
fill_pattern = PatternFill(patternType='solid', fgColor='B60C50')

# color fill the cell B5
ws['B5'].fill = fill_pattern


wb.save('Example.xlsx')

