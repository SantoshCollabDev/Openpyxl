import openpyxl

wb = openpyxl.load_workbook("Example.xlsx")


# To fetch value 400
ws = wb['Balance']
print(ws['B5'].value)    # remeber --->  .value


# To fetch value Robin using cell corodinates 
print(ws.cell(row=6, column=1).value)


# To fetch range of values rows from 1 - 5 & both column values
values_in_range = ws['A2':'B5']
print(values_in_range)  # values in the form of tuples
# UNPACK the TUPLES & read values
for row,col in values_in_range:
  print(row.value, col.value)  # for each tuple print row & col values


wb.save