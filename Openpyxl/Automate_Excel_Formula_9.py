import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook("Example.xlsx")
ws = wb['Balance']

ws['B10'] = "=SUM(B2:B8)"
ws['B11'] = "=AVERAGE(B2:B8)"

ws['E1'] = "Balance after a year"
ws['E1'].font = Font(bold=True, name='Arial', size=10)

for i in range(2,9):
  # ( balance * interest rate ) + balance  i.e. (b2*c2) + b2
  balance = ws.cell(row=i,column=2).value
  intrate = ws.cell(row=i, column=4).value
  final_balance = (balance * intrate) + balance
  ws.cell(row=i,column=5).value = final_balance

wb.save("Example.xlsx")