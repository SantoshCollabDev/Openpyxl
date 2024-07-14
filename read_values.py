import openpyxl

# load excel to wb object
wb = openpyxl.load_workbook("example.xlsx")

# get control on specific work sheet

ws = wb['balance']

# read balance value of customer 'D'
print(ws['B5'].value)

# access value using ---- ws.cell() method
print(ws.cell(row=5, column=2).value)

# access range of values
values_range = ws['A2':'B8']
print(values_range)
# output tuples
# ((<Cell 'balance'.A2>, <Cell 'balance'.B2>), 
#  (<Cell 'balance'.A3>, <Cell 'balance'.B3>), 
#  (<Cell 'balance'.A4>, <Cell 'balance'.B4>), 
#  (<Cell 'balance'.A5>, <Cell 'balance'.B5>), 
#  (<Cell 'balance'.A6>, <Cell 'balance'.B6>), 
#  (<Cell 'balance'.A7>, <Cell 'balance'.B7>), 
#  (<Cell 'balance'.A8>, <Cell 'balance'.B8>))

# do tuple unpacking to read values

for a, b in values_range:
  print(a.value,b.value)