import openpyxl

daily_data = openpyxl.load_workbook("Daily_Sheet.xlsx")
master_data = openpyxl.load_workbook("Master_Sheet.xlsx")

daily_sheet = daily_data['ddata']
master_sheet = master_data['mdata']

for i in daily_sheet.iter_rows():
  # print(i[0].value)
  id = i[0].value
  row_number = i[0].row
  # print(row_number)
  for j in master_sheet.iter_rows():
    if j[0].value == id:
      daily_sheet.cell(row=row_number,column=4).value = j[1].value
      daily_sheet.cell(row=row_number,column=5).value = j[2].value
      daily_sheet.cell(row=row_number,column=6).value = j[3].value

daily_data.save("Daily_Sheet_updated.xlsx")

