import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

ws = wb['balance']

# refer existing value
# print(ws['B8'].value)
# --------------------------------------
# UDPATE EXISTING VALUE
# ws['B8'].value += 100
# print(ws['B8'].value)
# --------------------------------------
# ADD NEW VALUE
# ws['A9'].value = 'H'
# ws['B9'].value = 8000
# --------------------------------------
# refer existing value --- USING CELL method
# print(ws.cell(row=9,column=2).value)
# --------------------------------------
# UPDATE using CELL emthod
# ws.cell(row=10,column=1).value = 'I'
# ws.cell(row=10,column=2).value = 9000
# print(ws.cell(row=10,column=1).value)
# print(ws.cell(row=10, column=2).value)
# --------------------------------------

ws['C1'] = "Double Balance"

for i in range(2,11):
  c_col = ws.cell(row=i,column=2).value
  c_value = c_col * 2
  # print(c_value)
  ws.cell(row=i, column=3).value = c_value

wb.save("example.xlsx")