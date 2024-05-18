import datetime
from datetime import datetime

import openpyxl

wb = openpyxl.load_workbook("check_sort.xlsx")
ws = wb.active

excel_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
  excel_data.append(row)

excel_data.sort(key=lambda row: datetime.strptime(row[2], '%m/%d/%Y'))
print(excel_data[0:5])

ws.delete_rows(2,ws.max_row-1)

for row in excel_data:
  ws.append(row)

wb.save("sorted_sheet.xlsx")

