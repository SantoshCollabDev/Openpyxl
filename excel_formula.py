import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook("example.xlsx")

ws = wb['balance']

# ws['B12'] = "=SUM(B2:B10)"
# ws['B13'] = "=AVERAGE(B2:B10)"

ws['E1'] = 'Total Balance After 1 Yr'
ws['E1'].font = Font(bold=True)

for i in range(2,11):
  balance = ws.cell(i,3).value
  interest = ws.cell(i,4).value
  final_value = balance + ( balance * interest )
  ws.cell(i,5).value = final_value

wb.save("example.xlsx")
