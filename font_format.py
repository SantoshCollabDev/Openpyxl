import openpyxl
from openpyxl.styles import Font, Color

wb = openpyxl.load_workbook("example.xlsx")

ws = wb["balance"]

font_style = Font(name="Calibri", 
                  size=11, 
                  color='F30E0E',
                  bold=True)

c1 = ws['C1']
c1.font = font_style 

for i in range(2,11):
  ws.cell(row=i,column=3).font = font_style

wb.save("example.xlsx")

