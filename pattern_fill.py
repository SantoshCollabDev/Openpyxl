import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("example.xlsx")

ws = wb['balance']

fill_pattern = PatternFill(patternType='solid',
                           fgColor='F30E0E')

ws['B4'].fill = fill_pattern


wb.save("example.xlsx")